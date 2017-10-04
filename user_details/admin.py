import openpyxl
import datetime
from datetime import timedelta
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from .models import  UserDetails


class UserDetailsAdmin(admin.ModelAdmin):
    list_display = ['person', 'display_photo']
    list_filter = ['state', 'person', 'created_at']

    def display_photo(self, obj):
        return '<img src="/static/{}" height="150", width="150" alt="Image">'.format(obj.image.name.split("/")[-1])

    actions = ['Report_on_User_added_to_Database']
    def Report_on_User_added_to_Database(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Report on User added to Database"
        start_of_week = datetime.datetime.today() - timedelta(days=datetime.datetime.today().weekday())
        userdetails = UserDetails.objects.filter(created_at__gte=start_of_week).values_list('person__username', flat=True)
        write_to_xlsx(1, userdetails, "WEEK", ws)
        start_of_month = datetime.date.today().replace(day=1)
        userdetails = UserDetails.objects.filter(created_at__gte=start_of_month).values_list('person__username', flat=True)
        write_to_xlsx(2, userdetails, "MONTH", ws)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=user_added_to_db_current_week.xlsx'
        wb.save(response)
        return response

    display_photo.allow_tags = True
admin.site.register(UserDetails, UserDetailsAdmin)


def write_to_xlsx(col, users_list, cal_title, workseet):
    row = 1
    cell = workseet.cell(row=row, column=col)
    cell.value = cal_title
    for user in users_list:
        row = row + 1
        c = workseet.cell(row=row, column=col)
        c.value = str(user)